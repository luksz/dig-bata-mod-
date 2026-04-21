#!/usr/bin/env python3
"""Generate validation artifacts and supporting figures."""

import argparse
import csv
import os
import re
import subprocess
import textwrap
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

os.environ.setdefault("MPLCONFIGDIR", str(Path.cwd() / ".matplotlib"))

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


PRICE_PER_SQM_LIMIT = 4725.0
DIGIT_TO_TOWN = {
    0: "BEDOK",
    1: "BUKIT PANJANG",
    2: "CLEMENTI",
    3: "CHOA CHU KANG",
    4: "HOUGANG",
    5: "JURONG WEST",
    6: "PASIR RIS",
    7: "TAMPINES",
    8: "WOODLANDS",
    9: "YISHUN",
}
SELECTED_PAIRS = [(1, 80), (1, 147), (3, 120), (8, 150)]


@dataclass(frozen=True)
class QuerySpec:
    matric_number: str
    start_year: int
    start_month: int
    towns: Tuple[str, ...]

    @property
    def start_key(self) -> int:
        return self.start_year * 12 + self.start_month


@dataclass(frozen=True)
class RawRow:
    month_text: str
    year: int
    month_num: int
    month_key: int
    town: str
    block: str
    floor_area: float
    flat_model: str
    lease_commence_date: str
    resale_price: float
    price_per_sqm: float


def round_half_up(value: float) -> int:
    return int(value + 0.5)


def parse_month_text(raw_month: str) -> Tuple[int, int]:
    parsed = datetime.strptime(raw_month.strip(), "%b-%y")
    return parsed.year, parsed.month


def pair_from_text(text_value: str) -> Tuple[int, int]:
    return tuple(int(number) for number in re.findall(r"\d+", text_value))  # type: ignore[return-value]


def format_floor_area(value: float) -> str:
    if value.is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def build_query_spec(matric_number: str) -> QuerySpec:
    digits = [int(ch) for ch in matric_number if ch.isdigit()]
    if len(digits) < 2:
        raise ValueError("Matric number must contain at least two digits.")

    last_digit = digits[-1]
    second_last_digit = digits[-2]
    start_year = 2015 + ((last_digit - 5) % 10)
    start_month = 10 if second_last_digit == 0 else second_last_digit
    towns = tuple(sorted({DIGIT_TO_TOWN[digit] for digit in digits}))
    return QuerySpec(matric_number, start_year, start_month, towns)


def load_raw_rows(csv_path: Path) -> List[RawRow]:
    rows: List[RawRow] = []
    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            year, month_num = parse_month_text(row["month"])
            floor_area = float(row["floor_area_sqm"])
            resale_price = float(row["resale_price"])
            rows.append(
                RawRow(
                    month_text=row["month"],
                    year=year,
                    month_num=month_num,
                    month_key=year * 12 + month_num,
                    town=row["town"],
                    block=row["block"],
                    floor_area=floor_area,
                    flat_model=row["flat_model"],
                    lease_commence_date=row["lease_commence_date"],
                    resale_price=resale_price,
                    price_per_sqm=resale_price / floor_area,
                )
            )
    return rows


def load_output_rows(output_path: Path) -> List[Dict[str, str]]:
    with output_path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def row_key(row: RawRow) -> Tuple[float, int, str, str, str, str]:
    return (
        row.price_per_sqm,
        row.month_key,
        row.town,
        row.block,
        row.flat_model,
        row.lease_commence_date,
    )


def filtered_candidates(rows: Iterable[RawRow], spec: QuerySpec, x_value: int, y_value: int) -> List[RawRow]:
    end_key = spec.start_key + x_value - 1
    matched = [
        row
        for row in rows
        if spec.start_key <= row.month_key <= end_key and row.town in spec.towns and row.floor_area >= y_value
    ]
    return sorted(matched, key=row_key)


def write_validation_summary_csv(summary_rows: Sequence[Dict[str, str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "Pair",
                "Candidate_Count",
                "Independent_Min_Raw",
                "Independent_Min_Rounded",
                "Program_Min_Rounded",
                "Matched",
                "Year",
                "Month",
                "Town",
                "Block",
                "Floor_Area",
                "Flat_Model",
                "Lease_Commence_Date",
            ],
        )
        writer.writeheader()
        writer.writerows(summary_rows)


def write_validation_workbook(
    summary_rows: Sequence[Dict[str, str]],
    candidate_map: Dict[Tuple[int, int], List[RawRow]],
    workbook_path: Path,
) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_headers = [
        "Pair",
        "Candidate Count",
        "Independent Raw Min",
        "Independent Rounded Min",
        "Program Rounded Min",
        "Match",
        "Year",
        "Month",
        "Town",
        "Block",
        "Floor Area",
        "Flat Model",
        "Lease Commence Date",
    ]
    summary_sheet.append(summary_headers)
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row in summary_rows:
        summary_sheet.append(
            [
                row["Pair"],
                int(row["Candidate_Count"]),
                float(row["Independent_Min_Raw"]),
                int(row["Independent_Min_Rounded"]),
                int(row["Program_Min_Rounded"]),
                row["Matched"],
                row["Year"],
                row["Month"],
                row["Town"],
                row["Block"],
                row["Floor_Area"],
                row["Flat_Model"],
                row["Lease_Commence_Date"],
            ]
        )

    for column in "ABCDEFGHIJKLM":
        summary_sheet.column_dimensions[column].width = 18

    for pair, rows in candidate_map.items():
        pair_name = f"Q_{pair[0]}_{pair[1]}"
        sheet = workbook.create_sheet(title=pair_name)
        header_row = 7
        data_start_row = 8

        sheet["A1"] = "Pair"
        sheet["B1"] = f"({pair[0]}, {pair[1]})"
        sheet["A2"] = "Candidate count"
        sheet["B2"] = len(rows)
        sheet["A3"] = "Excel MIN formula"
        sheet["B3"] = f"=MIN(H{data_start_row}:H{data_start_row + len(rows) - 1})"
        sheet["A4"] = "Rounded minimum"
        sheet["B4"] = "=ROUND(B3,0)"
        sheet["A5"] = "Program rounded output"
        matched_row = next(summary for summary in summary_rows if summary["Pair"] == f"({pair[0]}, {pair[1]})")
        sheet["B5"] = int(matched_row["Program_Min_Rounded"])
        sheet["A6"] = "Match"
        sheet["B6"] = '=IF(B4=B5,"YES","NO")'

        for label_cell in ("A1", "A2", "A3", "A4", "A5", "A6"):
            sheet[label_cell].font = Font(bold=True)

        headers = [
            "Month",
            "Town",
            "Block",
            "Floor_Area",
            "Flat_Model",
            "Lease_Commence_Date",
            "Resale_Price",
            "Price_Per_Sqm_Raw",
            "Price_Per_Sqm_Rounded",
        ]
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=header_row, column=column_index, value=header)
        for cell in sheet[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            sheet.append(
                [
                    row.month_text,
                    row.town,
                    row.block,
                    row.floor_area,
                    row.flat_model,
                    row.lease_commence_date,
                    row.resale_price,
                    row.price_per_sqm,
                    round_half_up(row.price_per_sqm),
                ]
            )

        for column in "ABCDEFGHI":
            sheet.column_dimensions[column].width = 18

        sheet.freeze_panes = f"A{data_start_row}"

    workbook.save(workbook_path)


def write_final_checks(
    output_rows: Sequence[Dict[str, str]],
    output_path: Path,
    spec: QuerySpec,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    header_ok = list(output_rows[0].keys()) == [
        "(x, y)",
        "Year",
        "Month",
        "Town",
        "Block",
        "Floor_Area",
        "Flat_Model",
        "Lease_Commence_Date",
        "Price_Per_Square_Meter",
    ]
    sorted_ok = all(
        pair_from_text(output_rows[idx]["(x, y)"]) < pair_from_text(output_rows[idx + 1]["(x, y)"])
        for idx in range(len(output_rows) - 1)
    )
    limit_ok = all(int(row["Price_Per_Square_Meter"]) <= int(PRICE_PER_SQM_LIMIT) for row in output_rows)
    blank_ok = all(all(value != "" for value in row.values()) for row in output_rows)

    lines = [
        f"Matric number: {spec.matric_number}",
        f"Derived start window: {spec.start_year}-{spec.start_month:02d}",
        f"Matched towns: {', '.join(spec.towns)}",
        f"Output row count: {len(output_rows)}",
        f"Header check: {'PASS' if header_ok else 'FAIL'}",
        f"Sorted by (x, y): {'PASS' if sorted_ok else 'FAIL'}",
        f"All price-per-sqm values <= 4725: {'PASS' if limit_ok else 'FAIL'}",
        f"No blank fields in emitted rows: {'PASS' if blank_ok else 'FAIL'}",
    ]
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def wrapped(text_value: str, width: int = 95) -> str:
    return "\n".join(textwrap.wrap(text_value, width=width))


def save_text_figure(text_value: str, output_path: Path, title: str, figsize: Tuple[float, float] = (10, 3.4)) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig = plt.figure(figsize=figsize)
    fig.patch.set_facecolor("white")
    fig.text(0.03, 0.94, title, fontsize=14, weight="bold", va="top")
    fig.text(
        0.03,
        0.86,
        text_value,
        family="monospace",
        fontsize=10,
        va="top",
        bbox={"boxstyle": "round", "facecolor": "#F5F5F5", "edgecolor": "#C8C8C8"},
    )
    plt.axis("off")
    plt.savefig(output_path, bbox_inches="tight", dpi=180)
    plt.close(fig)


def save_table_figure(
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
    output_path: Path,
    title: str,
    note: str = "",
    figsize: Tuple[float, float] = (11.0, 3.4),
    font_size: int = 9,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=figsize)
    ax.axis("off")
    table = ax.table(
        cellText=rows,
        colLabels=headers,
        cellLoc="center",
        loc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(font_size)
    table.scale(1, 1.35)
    for (row_index, col_index), cell in table.get_celld().items():
        if row_index == 0:
            cell.set_facecolor("#D9EAF7")
            cell.set_text_props(weight="bold")
        elif row_index % 2 == 1:
            cell.set_facecolor("#F8FBFD")
    fig.suptitle(title, fontsize=14, fontweight="bold", y=0.98)
    if note:
        fig.text(0.02, 0.03, wrapped(note, width=120), fontsize=9, va="bottom")
    plt.savefig(output_path, bbox_inches="tight", dpi=180)
    plt.close(fig)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate validation artifacts and supporting figures.")
    parser.add_argument("matric_number", help="Matriculation number used for the project.")
    parser.add_argument("--input", default="ResalePricesSingapore.csv", help="Input resale CSV file.")
    parser.add_argument(
        "--output",
        default=None,
        help="Output scan result CSV. Defaults to ScanResult_<MatricNum>.csv.",
    )
    args = parser.parse_args()

    repo_root = Path.cwd()
    csv_path = repo_root / args.input
    output_path = repo_root / (args.output or f"ScanResult_{args.matric_number}.csv")
    spec = build_query_spec(args.matric_number)

    raw_rows = load_raw_rows(csv_path)
    output_rows = load_output_rows(output_path)
    output_map = {pair_from_text(row["(x, y)"]): row for row in output_rows}

    validation_dir = repo_root / "validation"
    figure_dir = repo_root / "report_assets" / "figures"

    summary_rows: List[Dict[str, str]] = []
    candidate_map: Dict[Tuple[int, int], List[RawRow]] = {}

    for pair in SELECTED_PAIRS:
        candidates = filtered_candidates(raw_rows, spec, pair[0], pair[1])
        candidate_map[pair] = candidates
        best = candidates[0]
        program = output_map[pair]
        matched = (
            str(best.year) == program["Year"]
            and f"{best.month_num:02d}" == program["Month"]
            and best.town == program["Town"]
            and best.block == program["Block"]
            and format_floor_area(best.floor_area) == program["Floor_Area"]
            and best.flat_model == program["Flat_Model"]
            and best.lease_commence_date == program["Lease_Commence_Date"]
            and str(round_half_up(best.price_per_sqm)) == program["Price_Per_Square_Meter"]
        )
        summary_rows.append(
            {
                "Pair": f"({pair[0]}, {pair[1]})",
                "Candidate_Count": str(len(candidates)),
                "Independent_Min_Raw": f"{best.price_per_sqm:.6f}",
                "Independent_Min_Rounded": str(round_half_up(best.price_per_sqm)),
                "Program_Min_Rounded": program["Price_Per_Square_Meter"],
                "Matched": "YES" if matched else "NO",
                "Year": str(best.year),
                "Month": f"{best.month_num:02d}",
                "Town": best.town,
                "Block": best.block,
                "Floor_Area": format_floor_area(best.floor_area),
                "Flat_Model": best.flat_model,
                "Lease_Commence_Date": best.lease_commence_date,
            }
        )

    write_validation_summary_csv(summary_rows, validation_dir / f"{args.matric_number}_validation_summary.csv")
    write_validation_workbook(summary_rows, candidate_map, validation_dir / f"{args.matric_number}_validation_workbook.xlsx")
    write_final_checks(output_rows, validation_dir / "final_checks.txt", spec)

    run_output = subprocess.run(
        ["python3", "source/scan_resale_prices.py", args.matric_number],
        cwd=repo_root,
        check=True,
        capture_output=True,
        text=True,
    )
    run_text = f"$ python3 source/scan_resale_prices.py {args.matric_number}\n{run_output.stdout.strip()}"
    save_text_figure(run_text, figure_dir / "run_output.png", "Program Execution")

    preview_rows = [list(output_rows[idx].values()) for idx in [0, 1, 2, 3, len(output_rows) - 4, len(output_rows) - 3, len(output_rows) - 2, len(output_rows) - 1]]
    preview_rows.insert(4, ["...", "...", "...", "...", "...", "...", "...", "...", "..."])
    save_table_figure(
        headers=list(output_rows[0].keys()),
        rows=preview_rows,
        output_path=figure_dir / "output_preview.png",
        title="Output Preview",
        note="Top four and bottom four rows from ScanResult_U2221027H.csv. The preview shows the required field order and the expected sorting by (x, y).",
        figsize=(12.0, 4.2),
        font_size=7,
    )

    validation_summary_rows = [
        [
            row["Pair"],
            row["Candidate_Count"],
            row["Independent_Min_Rounded"],
            row["Program_Min_Rounded"],
            row["Matched"],
            row["Town"],
            row["Block"],
            f"{row['Year']}-{row['Month']}",
        ]
        for row in summary_rows
    ]
    save_table_figure(
        headers=[
            "Pair",
            "Candidates",
            "Independent",
            "Program",
            "Match",
            "Town",
            "Block",
            "Month",
        ],
        rows=validation_summary_rows,
        output_path=figure_dir / "validation_summary.png",
        title="Independent Validation Summary",
        note="Each pair was recomputed from the raw CSV with a separate brute-force pass.",
        figsize=(11.0, 2.8),
        font_size=9,
    )

    for pair in ((1, 80), (1, 147)):
        candidates = candidate_map[pair][:6]
        candidate_rows = [
            [
                row.month_text,
                row.town,
                row.block,
                format_floor_area(row.floor_area),
                row.flat_model,
                f"{row.price_per_sqm:.3f}",
                str(round_half_up(row.price_per_sqm)),
            ]
            for row in candidates
        ]
        best = candidates[0]
        note = (
            f"Pair ({pair[0]}, {pair[1]}): the sheet in the validation workbook uses Excel formulas "
            f"=MIN(...) and =ROUND(...,0). The minimum raw value is {best.price_per_sqm:.6f}, which rounds to {round_half_up(best.price_per_sqm)}."
        )
        save_table_figure(
            headers=["Month", "Town", "Block", "Floor_Area", "Flat_Model", "Raw PPSQM", "Rounded"],
            rows=candidate_rows,
            output_path=figure_dir / f"validation_pair_{pair[0]}_{pair[1]}.png",
            title=f"Validation Example for ({pair[0]}, {pair[1]})",
            note=note,
            figsize=(11.0, 2.8),
            font_size=8,
        )

    print(f"Wrote validation summary to {validation_dir / f'{args.matric_number}_validation_summary.csv'}")
    print(f"Wrote validation workbook to {validation_dir / f'{args.matric_number}_validation_workbook.xlsx'}")


if __name__ == "__main__":
    main()
